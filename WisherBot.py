from telethon.sync import TelegramClient
import openpyxl as xl
from datetime import date, datetime

api_id = '13516683'
api_hash = '28436249d6b2b2ec6fa68eb5e30a99d0'
token = '5029332615:AAEd7rjvfkCtMOz9AK39QDdqDLS4TrCRAzw'
username = 'WisherB'

time = datetime.now()
req_format = time.strftime("%d/%m")
wb = xl.load_workbook('Employees_birthday.xlsx')
ws = wb['Sheet1']
Birthday_persons = []

for row in range(2,ws.max_row+1):
    if(ws.cell(row,3).value[0:5] == req_format):
        Birthday_persons.append(ws.cell(row,2).value)
      
Birthday_persons = ','.join(Birthday_persons)
message = f"Hii {Birthday_persons} \nWishing you many more happy returns of the day"
print(message)

bot = TelegramClient('bot', api_id , api_hash).start(bot_token=token)
group = bot.get_entity(username)
# bot.send_message(group,message)
bot.send_file(group,file = open('BirthdayImg.jfif','rb'),caption=message)


